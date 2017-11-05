from abc import ABCMeta, abstractmethod
import sys
import csv
import openpyxl


class FileReader(metaclass=ABCMeta):
    def read_file(self, filename, validator):
        file = self.open_file(filename)
        if file:
            data = self.parse_file(file)
            valid_data = self.check_data(data, validator)
            if valid_data:
                return valid_data
            else:
                return False
        else:
            return False

    def open_file(self, filename):
        try:
            return open(filename, "r")
        except FileNotFoundError:
            print('The file was not found', file=sys.stderr)
            return False

    @abstractmethod
    def parse_file(self, file):
        pass

    def check_data(self, data, validator):
        valid_data = []
        for line in data:
            if validator.check_line(line):
                valid_data.append(line)
            else:
                print('Entry failed validation', file=sys.stderr)
        if validator.check_data_set(valid_data):

            return valid_data
        else:
            print("There were no valid entries in the file", file=sys.stderr)
            return False


class CSVReader(FileReader):
    def parse_file(self, file):
        reader = csv.DictReader(file, delimiter=',')
        the_list = []
        for line in reader:
            employee = dict()
            employee["EMPID"] = str(line["emp_id"])
            employee["GENDER"] = str(line["gender"])
            employee["AGE"] = str(line["age"])
            employee["SALES"] = str(line["sales"])
            employee["BMI"] = str(line["bmi"])
            employee["SALARY"] = str(line["salary"])
            employee["BIRTHDAY"] = str(line["birthday"])
            the_list.append(employee)
        return the_list


class TXTReader(FileReader):
    def parse_file(self, file):
        the_list = []
        for line in file:
            dictionary = {}
            entries = line.split(";")
            for entry in entries:
                if len(entry.split("=")) == 2:
                    key = entry.split("=")[0]
                    value = entry.split("=")[1]
                    value = value.rstrip('\n')
                    dictionary[key] = value
            the_list.append(dictionary)
        return the_list


class XLSXReader(FileReader):
    def open_file(self, filename):
        try:
            return openpyxl.load_workbook(filename)
        except FileNotFoundError:
            print("File not found!", file=sys.stderr)
            return False

    def parse_file(self, file):
        sheet = file.active
        the_list = []
        for x in range(1, 29):
            employee = dict()
            employee["EMPID"] = str(sheet.cell(column=1, row=x).value)
            employee["GENDER"] = str(sheet.cell(column=2, row=x).value)
            employee["AGE"] = str(sheet.cell(column=3, row=x).value)
            employee["SALES"] = str(sheet.cell(column=4, row=x).value)
            employee["BMI"] = str(sheet.cell(column=5, row=x).value)
            employee["SALARY"] = str(sheet.cell(column=6, row=x).value)
            employee["BIRTHDAY"] = str(sheet.cell(column=7, row=x).value)
            the_list.append(employee)
        return the_list


if __name__ == "__main__":
    import doctest
    doctest.testmod(verbose=1)
